"""Generate and save sample Excel templates for each import module.

Each template contains the exact column headers the Import Wizard
expects, plus 2-3 example rows so users can see the expected format.
Templates are written to a temporary directory and can be saved via
a file-dialog in the UI.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _style_worksheet(ws, headers: list[str], data: list[list], notes: list[str] | None = None) -> None:
    """Apply professional styling to a worksheet."""
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    data_font = Font(name="Calibri", size=11, color="374151")
    data_fill_even = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    note_font = Font(name="Calibri", size=10, italic=True, color="6B7280")

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write sample data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = data_fill_even

    # Write notes below data
    if notes:
        note_row = len(data) + 3
        for i, note in enumerate(notes):
            cell = ws.cell(row=note_row + i, column=1, value=note)
            cell.font = note_font
            ws.merge_cells(
                start_row=note_row + i,
                start_column=1,
                end_row=note_row + i,
                end_column=min(len(headers), 6),
            )

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 35)

    # Freeze top row
    ws.freeze_panes = "A2"


def generate_tenders_template(path: str | Path) -> Path:
    """Create a sample Tenders import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tenders"

    headers = [
        "firm_name",
        "bid_no",
        "organisation",
        "department",
        "state",
        "location",
        "publish_date",
        "due_date",
        "tender_value",
        "emd",
        "participation_status",
        "contract_period_months",
        "quantity",
        "nature_of_work",
        "category",
        "service_days",
        "publish_rate",
        "quoted_rates",
    ]

    data = [
        [
            "Acme Corp",
            "BID-2026-001",
            "Municipal Corporation",
            "Public Works",
            "Maharashtra",
            "Mumbai",
            "2026-01-15",
            "2026-02-15",
            500000,
            25000,
            "Participated",
            12,
            100,
            "Catering Service",
            "Food & Beverage",
            26,
            50.00,
            48.50,
        ],
        [
            "Acme Corp",
            "BID-2026-002",
            "State Transport Dept",
            "Transport",
            "Gujarat",
            "Ahmedabad",
            "2026-02-01",
            "2026-03-01",
            1200000,
            60000,
            "Not Participated",
            24,
            200,
            "Housekeeping",
            "Facility Management",
            30,
            200.00,
            195.00,
        ],
        [
            "Beta Solutions",
            "GEM/2026/B/789",
            "Indian Railways",
            "Catering",
            "Delhi",
            "New Delhi",
            "2026-03-10",
            "2026-04-10",
            800000,
            "Nil",
            "Participated in Support",
            6,
            150,
            "Pantry Car Service",
            "Railway Catering",
            180,
            88.89,
            85.00,
        ],
    ]

    notes = [
        "NOTES:",
        "• firm_name must match an existing firm in the database exactly.",
        "• Dates: use YYYY-MM-DD or DD/MM/YYYY format.",
        "• emd: can be a number or text like 'Nil', 'Exempt', 'N/A'.",
        "• participation_status: Participated, Participated in Support, Not Participated, Cancelled.",
        "• All numeric fields (tender_value, quantity, etc.) must be numbers.",
    ]

    _style_worksheet(ws, headers, data, notes)
    p = Path(path)
    wb.save(str(p))
    return p


def generate_compliance_template(path: str | Path) -> Path:
    """Create a sample Compliance import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"

    headers = [
        "firm_name",
        "document_name",
        "document_type",
        "certificate_no",
        "issue_date",
        "expiry_date",
        "status",
    ]

    data = [
        [
            "Acme Corp",
            "GST Registration",
            "Tax Registration",
            "29AABCT1234F1Z5",
            "2025-04-01",
            "2026-03-31",
            "Active",
        ],
        [
            "Acme Corp",
            "FSSAI License",
            "Food License",
            "12345678901234",
            "2025-06-15",
            "2026-06-14",
            "To Be Renewed",
        ],
        [
            "Beta Solutions",
            "Labour License",
            "Statutory",
            "LC/2025/MH/001",
            "2025-01-01",
            "2025-12-31",
            "Expired",
        ],
    ]

    notes = [
        "NOTES:",
        "• firm_name must match an existing firm in the database.",
        "• Dates: use YYYY-MM-DD or DD/MM/YYYY format.",
        "• status: Active, To Be Renewed, Under Renewal, Expired, Not Applicable.",
        "• document_name is required.",
    ]

    _style_worksheet(ws, headers, data, notes)
    p = Path(path)
    wb.save(str(p))
    return p


def generate_estamps_template(path: str | Path) -> Path:
    """Create a sample E-Stamps import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "E-Stamps"

    headers = [
        "firm_name",
        "entry_date",
        "tender_name_text",
        "quantity",
        "unit_rate",
    ]

    data = [
        ["Acme Corp", "2026-01-10", "BID-2026-001 Municipal Corp", 2, 100],
        ["Acme Corp", "2026-02-05", "BID-2026-002 State Transport", 1, 500],
        ["Beta Solutions", "2026-03-15", "GEM/2026/B/789 Railways", 3, 100],
    ]

    notes = [
        "NOTES:",
        "• firm_name must match an existing firm in the database.",
        "• entry_date: use YYYY-MM-DD or DD/MM/YYYY format. Defaults to today if empty.",
        "• unit_rate: denomination value of the e-stamp (e.g. 100, 500, 1000).",
        "• quantity: number of e-stamps (must be a whole number).",
        "• tender_name_text: optional description/reference for the e-stamp.",
    ]

    _style_worksheet(ws, headers, data, notes)
    p = Path(path)
    wb.save(str(p))
    return p


def generate_vault_template(path: str | Path) -> Path:
    """Create a sample Password Vault import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Password Vault"

    headers = [
        "firm_name",
        "portal_name",
        "portal_url",
        "username",
        "password",
        "dsc_holder",
        "dsc_expiry",
        "registered_mobile",
        "registered_email",
        "notes",
    ]

    data = [
        [
            "Acme Corp",
            "GeM Portal",
            "https://gem.gov.in",
            "acme_gem_user",
            "SecureP@ss123",
            "Rajesh Kumar",
            "2027-06-30",
            "9876543210",
            "acme@example.com",
            "Primary bidding account",
        ],
        [
            "Acme Corp",
            "IREPS",
            "https://ireps.gov.in",
            "acme_ireps",
            "Ir3p$Pass!",
            "Priya Sharma",
            "2026-12-31",
            "9876543211",
            "ireps@acme.com",
            "Railway e-procurement",
        ],
        [
            "Beta Solutions",
            "eTender Portal",
            "https://etenders.gov.in",
            "beta_etender",
            "B3ta$ecure#",
            "Amit Verma",
            "2027-03-15",
            "8765432109",
            "beta@solutions.com",
            "State tenders",
        ],
    ]

    notes = [
        "NOTES:",
        "• firm_name must match an existing firm in the database.",
        "• portal_name is required.",
        "• password & username will be encrypted (AES-256-GCM) before storage.",
        "• dsc_expiry: DSC expiry date in YYYY-MM-DD or DD/MM/YYYY format.",
        "• ⚠ WARNING: Delete this file after import — it contains plaintext passwords!",
    ]

    _style_worksheet(ws, headers, data, notes)
    p = Path(path)
    wb.save(str(p))
    return p


# Map module names (as used in the UI) to generator functions.
TEMPLATE_GENERATORS = {
    "Tenders": generate_tenders_template,
    "Compliance": generate_compliance_template,
    "E-Stamps": generate_estamps_template,
    "Password Vault": generate_vault_template,
}


def save_sample_template(module: str, dest_path: str | Path) -> Path:
    """Generate and save a sample template for the given module."""
    gen = TEMPLATE_GENERATORS.get(module)
    if gen is None:
        raise ValueError(f"Unknown module: {module}")
    return gen(dest_path)
