"""Excel Import Service."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from difflib import SequenceMatcher
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import inspect as sa_inspect
from sqlalchemy.orm import Session

from ..models.compliance import ComplianceDocument
from ..models.estamp import Estamp
from ..models.firm import Firm
from ..models.tender import Tender
from ..models.user import Role
from ..models.vault import VaultCredential
from .auth import create_user
from .tender_rates import computed_publish_rate_fields

# Text values commonly used in EMD column instead of numbers (India tenders).
_EMD_NON_NUMERIC = frozenset(
    {
        "nil",
        "n/a",
        "na",
        "-",
        "—",
        "exempt",
        "exemption",
        "not applicable",
    }
)


def _normalize_firm_name(name: str) -> str:
    """Collapse whitespace, strip common legal prefixes, casefold for matching."""
    s = unicodedata.normalize("NFKC", name)
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    # Remove leading M/s., Mr., etc. so Excel variants match DB names.
    s = re.sub(
        r"^(m/s\.?|mr\.?|mrs\.?|ms\.?|dr\.?|messrs\.?)\s+",
        "",
        s,
        flags=re.IGNORECASE,
    )
    s = re.sub(r"\s+", " ", s).strip()
    return s.casefold()


def _excel_scalar_to_str(val: Any) -> str:
    """Firm names from Excel may arrive as float (rare) or number."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _resolve_firm(session: Session, raw: Any) -> Firm | None:
    """Match firm by flexible name: normalized equality, then fuzzy (typos, prefixes)."""
    name = _excel_scalar_to_str(raw)
    if not name:
        return None

    firms = session.query(Firm).filter(Firm.is_archived.is_(False)).all()
    if not firms:
        firms = session.query(Firm).all()

    target = _normalize_firm_name(name)

    # 1) Normalized equality (handles spacing, M/s., Mr., case)
    for f in firms:
        if _normalize_firm_name(f.name) == target:
            return f

    # 2) Plain case-insensitive full string
    nl = name.lower()
    for f in firms:
        if f.name.strip().lower() == nl:
            return f

    # 3) Fuzzy: single best match when clearly closest (e.g. Johnry vs Johnny)
    best: Firm | None = None
    best_score = 0.0
    second_score = 0.0
    for f in firms:
        fn = _normalize_firm_name(f.name)
        score = SequenceMatcher(None, target, fn).ratio()
        if score > best_score:
            second_score = best_score
            best_score = score
            best = f
        elif score > second_score:
            second_score = score

    if best is not None and best_score >= 0.86 and (best_score - second_score) >= 0.01:
        return best

    return None


def _only_model_columns(model_cls: type, data: dict[str, Any]) -> dict[str, Any]:
    mapper = sa_inspect(model_cls)
    cols = {c.key for c in mapper.column_attrs}
    return {k: v for k, v in data.items() if k in cols}


def _coerce_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # DD/MM/YYYY or DD-MM-YYYY (common in Indian sheets)
    for sep in ("/", "-", "."):
        m = re.match(r"^(\d{1,2})" + re.escape(sep) + r"(\d{1,2})" + re.escape(sep) + r"(\d{4})$", s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return date(y, mo, d)
            except ValueError:
                pass
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _coerce_optional_money(val: Any) -> float | None:
    """Parse money columns; treat exemption / nil text as NULL."""
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    low = s.lower().replace(",", "").replace("₹", "").strip()
    if low in _EMD_NON_NUMERIC:
        return None
    if "exemption" in low or low == "nil" or low in ("na", "n/a", "-"):
        return None
    try:
        return float(low)
    except ValueError:
        return None


def parse_excel(file_path: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse an excel file into headers and data dictionaries.

    Headers align 1:1 with columns: empty header cells keep column index
    (``__empty_N``) so values do not shift under the wrong field names.
    """
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_row = list(rows[0])
    ncols = len(header_row)
    headers: list[str] = []
    for i, cell in enumerate(header_row):
        if cell is None or (isinstance(cell, str) and not str(cell).strip()):
            headers.append(f"__empty_{i}")
        else:
            headers.append(str(cell).strip())

    data: list[dict[str, Any]] = []
    for row in rows[1:]:
        row_list = list(row) + [None] * (ncols - len(row))
        row_list = row_list[:ncols]
        row_dict: dict[str, Any] = {}
        is_empty = True
        for i, val in enumerate(row_list):
            if i >= len(headers):
                break
            if isinstance(val, datetime):
                val = val.date()
            row_dict[headers[i]] = val
            if val is not None and str(val).strip():
                is_empty = False
        if not is_empty:
            data.append(row_dict)

    return headers, data


def process_import(
    session: Session, module: str, mapping: dict[str, str], data: list[dict[str, Any]]
) -> tuple[int, list[str]]:
    """Convert mapped data into ORM models and save them.

    If any row fails validation or DB integrity constraints,
    the entire transaction is rolled back and errors are returned.
    """
    success_count = 0
    errors: list[str] = []

    for idx, row in enumerate(data):
        try:
            mapped_row = {db_f: row.get(ex_h) for db_f, ex_h in mapping.items() if ex_h}

            firm_name = mapped_row.pop("firm_name", None)
            firm_id = None
            if firm_name is not None and str(firm_name).strip() != "":
                firm = _resolve_firm(session, firm_name)
                if firm:
                    firm_id = firm.id
                else:
                    display = _excel_scalar_to_str(firm_name)
                    errors.append(
                        f"Row {idx + 1}: Firm '{display}' not found in DB. "
                        "Add the firm under Firms first, or fix spelling to match exactly."
                    )
                    continue

            if module == "Firms":
                if not mapped_row.get("name"):
                    errors.append(f"Row {idx + 1}: Legal name is required")
                    continue
                session.add(Firm(**_only_model_columns(Firm, mapped_row)))

            elif module == "Tenders":
                if not firm_id:
                    errors.append(f"Row {idx + 1}: 'firm_name' is required for Tenders")
                    continue
                mapped_row["firm_id"] = firm_id

                for dcol in ("publish_date", "due_date"):
                    if dcol in mapped_row:
                        mapped_row[dcol] = _coerce_date(mapped_row.get(dcol))

                # EMD often contains "Nil", "MSE Exemption", etc. — treat as NULL.
                if "emd" in mapped_row:
                    mapped_row["emd"] = _coerce_optional_money(mapped_row.get("emd"))

                row_error: str | None = None
                for num_col in (
                    "tender_value",
                    "publish_rate",
                    "quoted_rates",
                    "contract_period_months",
                    "quantity",
                    "service_days",
                ):
                    if num_col not in mapped_row:
                        continue
                    v = mapped_row.get(num_col)
                    if v is None or v == "":
                        mapped_row[num_col] = None
                        continue
                    if isinstance(v, (int, float)):
                        mapped_row[num_col] = float(v)
                        continue
                    s = str(v).strip().replace(",", "")
                    try:
                        mapped_row[num_col] = float(s)
                    except ValueError:
                        row_error = f"Row {idx + 1}: {num_col} must be numeric (got {v!r})."
                        break

                if row_error:
                    errors.append(row_error)
                    continue

                tender_kw = _only_model_columns(Tender, mapped_row)
                cpm = tender_kw.get("contract_period_months")
                period_fallback = (float(cpm) * 30.0) if cpm else None
                auto_pr = computed_publish_rate_fields(
                    tender_value=tender_kw.get("tender_value"),
                    quantity=tender_kw.get("quantity"),
                    nature_of_work=tender_kw.get("nature_of_work"),
                    category=tender_kw.get("category"),
                    contract_period_months=tender_kw.get("contract_period_months"),
                    service_days=tender_kw.get("service_days"),
                    period_in_days_fallback=period_fallback,
                )
                if auto_pr is not None:
                    tender_kw["publish_rate"] = auto_pr
                session.add(Tender(**tender_kw))

            elif module == "Users":
                username = mapped_row.get("username")
                full_name = mapped_row.get("full_name")
                role = mapped_row.get("role", Role.EDITOR.value)
                if not username or not full_name:
                    errors.append(f"Row {idx + 1}: Username and Full Name required")
                    continue

                create_user(
                    session,
                    username=str(username).strip(),
                    full_name=str(full_name).strip(),
                    password="TempPassword123!",
                    role=role,
                )

            elif module == "E-Stamps":
                if not firm_id:
                    errors.append(f"Row {idx + 1}: 'firm_name' is required for E-Stamps")
                    continue
                mapped_row["firm_id"] = firm_id
                if mapped_row.get("entry_date"):
                    mapped_row["entry_date"] = _coerce_date(mapped_row.get("entry_date"))
                if not mapped_row.get("entry_date"):
                    mapped_row["entry_date"] = date.today()

                if mapped_row.get("quantity") not in (None, ""):
                    try:
                        mapped_row["quantity"] = int(mapped_row["quantity"])
                    except (TypeError, ValueError) as e:
                        errors.append(f"Row {idx + 1}: quantity must be an integer ({e}).")
                        continue
                if mapped_row.get("unit_rate") not in (None, ""):
                    try:
                        mapped_row["unit_rate"] = float(mapped_row["unit_rate"])
                    except (TypeError, ValueError):
                        errors.append(
                            f"Row {idx + 1}: unit_rate must be numeric (got {mapped_row.get('unit_rate')!r})."
                        )
                        continue

                session.add(Estamp(**_only_model_columns(Estamp, mapped_row)))

            elif module == "Compliance":
                if not firm_id:
                    errors.append(f"Row {idx + 1}: 'firm_name' is required for Compliance")
                    continue
                mapped_row["firm_id"] = firm_id
                if not mapped_row.get("document_name"):
                    errors.append(f"Row {idx + 1}: Document name required")
                    continue

                for dcol in ("issue_date", "expiry_date", "renewal_due_date"):
                    if dcol in mapped_row:
                        mapped_row[dcol] = _coerce_date(mapped_row.get(dcol))

                session.add(ComplianceDocument(**_only_model_columns(ComplianceDocument, mapped_row)))

            elif module == "Password Vault":
                if not firm_id:
                    errors.append(f"Row {idx + 1}: 'firm_name' is required for Password Vault")
                    continue
                mapped_row["firm_id"] = firm_id
                if not mapped_row.get("portal_name"):
                    errors.append(f"Row {idx + 1}: portal_name is required")
                    continue

                if "dsc_expiry" in mapped_row:
                    mapped_row["dsc_expiry"] = _coerce_date(mapped_row.get("dsc_expiry"))

                # Sensitive fields are stored as plaintext bytes for now;
                # a full implementation would require the vault key to encrypt.
                username_val = mapped_row.pop("username", None)
                password_val = mapped_row.pop("password", None)

                cred_kw = _only_model_columns(VaultCredential, mapped_row)
                if username_val:
                    cred_kw["username_enc"] = str(username_val).encode("utf-8")
                if password_val:
                    cred_kw["password_enc"] = str(password_val).encode("utf-8")

                session.add(VaultCredential(**cred_kw))

            success_count += 1

        except Exception as e:
            if str(e) == "abort row":
                continue
            errors.append(f"Row {idx + 1}: {str(e)}")

    if not errors:
        session.commit()
    else:
        session.rollback()
        success_count = 0

    return success_count, errors
