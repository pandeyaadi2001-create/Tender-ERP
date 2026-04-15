"""Excel importer for migrating the legacy .xlsx trackers.

Spec §9 risk #6: "Build the Excel-importer in week 1". This module
reads the four source workbooks described in the spec and inserts
rows into the corresponding tables. Column headers are best-effort
matched against a dictionary so variants between firms (Mr. Johnny vs
AV Engineers) don't need separate code paths.
"""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from ..models.compliance import ComplianceDocument
from ..models.estamp import Estamp
from ..models.firm import Firm
from ..models.tender import Tender
from .validators import validate_compliance, validate_estamp, validate_tender


# Maps a lowercased, whitespace-collapsed header name → canonical attribute.
TENDER_HEADER_MAP = {
    "sn": None,  # auto
    "publish date": "publish_date",
    "due date": "due_date",
    "due time": "due_time",
    "bid no": "bid_no",
    "bid no.": "bid_no",
    "name of organisation": "organisation",
    "organisation": "organisation",
    "state": "state",
    "department": "department",
    "location": "location",
    "issuing authority": "issuing_authority",
    "contract period": "contract_period_months",
    "contract period (months)": "contract_period_months",
    "quantity": "quantity",
    "publish rate": "publish_rate",
    "tender value": "tender_value",
    "pbg %": "pbg_percent",
    "pbg period": "pbg_period_months",
    "emd": "emd",
    "payment mode": "payment_mode",
    "quoted rates": "quoted_rates",
    "participation status": "participation_status",
    "participation date": "participation_date",
    "nature of work": "nature_of_work",
    "scope of work": "scope_of_work",
    "technical status": "technical_status",
    "financial status": "financial_status",
    "our status": "our_status",
    "l1 rates": "l1_rates",
    "month of award": "month_of_award",
}


def _norm(h: Any) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _coerce(attr: str, value: Any) -> Any:
    if value is None or value == "":
        return None
    if attr.endswith("_date"):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return datetime.fromisoformat(str(value)).date()
        except ValueError:
            return None
    if attr.endswith("_time"):
        if isinstance(value, time):
            return value
        if isinstance(value, datetime):
            return value.time()
        return None
    if attr in (
        "contract_period_months",
        "quantity",
        "publish_rate",
        "tender_value",
        "pbg_percent",
        "pbg_period_months",
        "emd",
        "quoted_rates",
        "l1_rates",
    ):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    return str(value).strip()


def import_tenders_xlsx(
    session: Session, workbook_path: Path, firm: Firm, sheet: str | None = None
) -> int:
    """Import a tender tracker workbook for one firm. Returns row count."""
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [_norm(h) for h in next(rows, ())]

    count = 0
    for row in rows:
        if not any(c not in (None, "") for c in row):
            continue
        payload: dict[str, Any] = {"firm_id": firm.id}
        for header, value in zip(headers, row):
            attr = TENDER_HEADER_MAP.get(header)
            if attr is None:
                continue
            payload[attr] = _coerce(attr, value)
        errors = validate_tender(payload)
        if errors:
            # Skip rows that fail hard validation; real migrator should
            # surface these in a UI report. Keeping it silent for v0.5.
            continue
        session.add(Tender(**{k: v for k, v in payload.items() if hasattr(Tender, k)}))
        count += 1
    session.flush()
    return count


COMPLIANCE_HEADER_MAP = {
    "certificate no": "certificate_no",
    "certificate no.": "certificate_no",
    "document type": "document_type",
    "document name": "document_name",
    "issuing authority": "issuing_authority",
    "issue date": "issue_date",
    "expiry date": "expiry_date",
    "expiry/assessment date": "expiry_date",
    "renewal due date": "renewal_due_date",
    "renewal status": "status",
    "responsible person": "responsible_person",
    "notes": "notes",
}


def import_compliance_xlsx(
    session: Session, workbook_path: Path, firm: Firm, sheet: str | None = None
) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [_norm(h) for h in next(rows, ())]
    count = 0
    for row in rows:
        if not any(c not in (None, "") for c in row):
            continue
        payload: dict[str, Any] = {"firm_id": firm.id}
        for header, value in zip(headers, row):
            attr = COMPLIANCE_HEADER_MAP.get(header)
            if attr is None:
                continue
            payload[attr] = _coerce(attr, value)
        if not payload.get("document_name"):
            continue
        errors = validate_compliance(payload)
        if errors:
            continue
        session.add(
            ComplianceDocument(
                **{k: v for k, v in payload.items() if hasattr(ComplianceDocument, k)}
            )
        )
        count += 1
    session.flush()
    return count


def import_estamps_xlsx(
    session: Session, workbook_path: Path, firm: Firm, sheet: str | None = None
) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [_norm(h) for h in next(rows, ())]
    count = 0
    for row in rows:
        if not any(c not in (None, "") for c in row):
            continue
        mapping: dict[str, Any] = {"firm_id": firm.id}
        for header, value in zip(headers, row):
            if header == "date":
                mapping["entry_date"] = _coerce("entry_date", value)
            elif header in ("tender name", "name of tender"):
                mapping["tender_name_text"] = _coerce("tender_name_text", value)
            elif header in ("no. of estamps", "quantity", "no of estamps"):
                try:
                    mapping["quantity"] = int(value)
                except (TypeError, ValueError):
                    pass
            elif header in ("unit rate", "rate"):
                try:
                    mapping["unit_rate"] = float(value)
                except (TypeError, ValueError):
                    pass
        if "entry_date" not in mapping or mapping.get("entry_date") is None:
            continue
        errors = validate_estamp(mapping)
        if errors:
            continue
        session.add(Estamp(**{k: v for k, v in mapping.items() if hasattr(Estamp, k)}))
        count += 1
    session.flush()
    return count
